# -*- coding: utf-8 -*-
"""
Created on Thu Jan  8 11:59:51 2026

@author: mg.kouame
"""

"""
✅ Lecture intelligente : Charge votre fichier Excel et détecte automatiquement les trimestres disponibles
✅ Menu interactif : Affiche un menu numéroté pour sélectionner facilement le trimestre
✅ Vérification des dépendances : Analyse automatiquement si les tâches respectent l'ordre d'exécution et affiche des alertes en cas de chevauchement
✅ Diagramme professionnel avec :

Barres bleues pour les tâches actives (symbole ■)
Weekends grisés automatiquement
Colonnes d'information (Ordre, Tâche, Début, Fin, Durée en jours)
Une colonne par jour avec format jj/mm
En-têtes en bleu foncé avec texte blanc
Bordures sur toutes les cellules
Volets figés pour faciliter la navigation
Titre du diagramme avec le trimestre

✅ Fichier de sortie : Créé dans le même dossier que votre fichier source avec le nom Diagramme_Gantt_T1_2026.xlsx
Comment utiliser le script :

Enregistrez le code dans un fichier nommé generer_gantt.py dans le dossier D:\ENEM_Working\Apurement salaire\Doosier_travail_Elodie\
Ouvrez l'invite de commandes dans ce dossier (Shift + clic droit > "Ouvrir dans le Terminal")
Exécutez :

bash   python generer_gantt.py

Suivez les instructions : Sélectionnez votre trimestre dans le menu

Le diagramme sera créé automatiquement ! 🎉
Voulez-vous que j'ajoute des fonctionnalités supplémentaires (graphique Excel, légende, etc.) ?Claude est une IA et peut faire des erreurs. Veuillez vérifier les réponses.

"""
"""
pip install pandas
pip install openpyxl
pip install xlsxwriter
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime, timedelta
import os

class GanttGenerator:
    def __init__(self, fichier_source):
        self.fichier_source = fichier_source
        self.df = None
        self.trimestre_choisi = None
        
    def lire_donnees(self):
        """Lit les données du fichier Excel source"""
        try:
            self.df = pd.read_excel(self.fichier_source)
            print("✓ Fichier chargé avec succès!")
            print(f"  Nombre de tâches trouvées : {len(self.df)}")
            return True
        except FileNotFoundError:
            print(f"✗ Erreur : Le fichier '{self.fichier_source}' est introuvable.")
            return False
        except Exception as e:
            print(f"✗ Erreur lors de la lecture du fichier : {e}")
            return False
    
    def afficher_menu_trimestre(self):
        """Affiche le menu de sélection du trimestre"""
        trimestres_disponibles = sorted(self.df['Timestre_a_debute'].unique())
        
        print("\n" + "="*50)
        print("   SÉLECTION DU TRIMESTRE")
        print("="*50)
        
        for idx, trimestre in enumerate(trimestres_disponibles, 1):
            print(f"  {idx}. {trimestre}")
        
        print("="*50)
        
        while True:
            try:
                choix = int(input("\nEntrez le numéro de votre choix : "))
                if 1 <= choix <= len(trimestres_disponibles):
                    self.trimestre_choisi = trimestres_disponibles[choix - 1]
                    print(f"\n✓ Trimestre sélectionné : {self.trimestre_choisi}")
                    return True
                else:
                    print(f"✗ Veuillez entrer un nombre entre 1 et {len(trimestres_disponibles)}")
            except ValueError:
                print("✗ Veuillez entrer un nombre valide")
    
    def filtrer_donnees(self):
        """Filtre les données selon le trimestre choisi"""
        self.df = self.df[self.df['Timestre_a_debute'] == self.trimestre_choisi].copy()
        self.df = self.df.sort_values('Ordre_exécution').reset_index(drop=True)
        
        # Convertir les dates en datetime
        self.df['Date_debut_Tache'] = pd.to_datetime(self.df['Date_debut_Tache'])
        self.df['Date_fin_tache'] = pd.to_datetime(self.df['Date_fin_tache'])
        
        print(f"✓ {len(self.df)} tâche(s) à afficher pour {self.trimestre_choisi}")
    
    def verifier_dependances(self):
        """Vérifie les dépendances entre tâches et affiche les alertes"""
        print("\n" + "-"*50)
        print("   VÉRIFICATION DES DÉPENDANCES")
        print("-"*50)
        
        problemes = []
        for i in range(1, len(self.df)):
            tache_precedente = self.df.iloc[i-1]
            tache_actuelle = self.df.iloc[i]
            
            if tache_actuelle['Date_debut_Tache'] < tache_precedente['Date_fin_tache']:
                problemes.append({
                    'tache': tache_actuelle['Intitule_taches'],
                    'ordre': tache_actuelle['Ordre_exécution'],
                    'debut': tache_actuelle['Date_debut_Tache'].strftime('%d/%m/%Y'),
                    'tache_prec': tache_precedente['Intitule_taches'],
                    'fin_prec': tache_precedente['Date_fin_tache'].strftime('%d/%m/%Y')
                })
        
        if problemes:
            print("⚠ Attention : Chevauchements détectés !")
            for p in problemes:
                print(f"  • Tâche #{p['ordre']} '{p['tache']}' commence le {p['debut']}")
                print(f"    avant la fin de '{p['tache_prec']}' (fin : {p['fin_prec']})")
        else:
            print("✓ Aucun chevauchement détecté - Les dépendances sont respectées")
        
        print("-"*50)
    
    def est_weekend(self, date):
        """Vérifie si une date tombe un weekend"""
        return date.weekday() >= 5  # 5=Samedi, 6=Dimanche
    
    def generer_diagramme(self):
        """Génère le diagramme de Gantt dans Excel"""
        date_debut_projet = self.df['Date_debut_Tache'].min()
        date_fin_projet = self.df['Date_fin_tache'].max()
        
        # Générer toutes les dates du projet
        dates = []
        date_courante = date_debut_projet
        while date_courante <= date_fin_projet:
            dates.append(date_courante)
            date_courante += timedelta(days=1)
        
        # Créer le DataFrame pour le Gantt
        gantt_data = []
        for _, tache in self.df.iterrows():
            ligne = {
                'Ordre': int(tache['Ordre_exécution']),
                'Tâche': tache['Intitule_taches'],
                'Début': tache['Date_debut_Tache'].strftime('%d/%m/%Y'),
                'Fin': tache['Date_fin_tache'].strftime('%d/%m/%Y'),
                'Durée (j)': (tache['Date_fin_tache'] - tache['Date_debut_Tache']).days + 1
            }
            
            # Ajouter une colonne pour chaque date
            for date in dates:
                if tache['Date_debut_Tache'] <= date <= tache['Date_fin_tache']:
                    ligne[date.strftime('%d/%m')] = '■'
                else:
                    ligne[date.strftime('%d/%m')] = ''
            
            gantt_data.append(ligne)
        
        gantt_df = pd.DataFrame(gantt_data)
        
        # Créer le fichier Excel
        nom_fichier = f"Diagramme_Gantt_{self.trimestre_choisi}.xlsx"
        chemin_sortie = os.path.join(os.path.dirname(self.fichier_source), nom_fichier)
        
        with pd.ExcelWriter(chemin_sortie, engine='openpyxl') as writer:
            gantt_df.to_excel(writer, sheet_name='Gantt Visuel', index=False)
            
            # Créer l'onglet récapitulatif
            self.creer_onglet_recap(writer)
        
        # Appliquer la mise en forme et créer le graphique
        self.appliquer_mise_en_forme(chemin_sortie, dates)
        self.creer_graphique_gantt(chemin_sortie)
        self.ajouter_legende(chemin_sortie)
        
        print(f"\n✓ Diagramme créé avec succès !")
        print(f"  Fichier : {nom_fichier}")
        print(f"  Emplacement : {os.path.dirname(self.fichier_source)}")
        
        return chemin_sortie
    
    def creer_onglet_recap(self, writer):
        """Crée un onglet récapitulatif avec les statistiques du projet"""
        stats_data = []
        
        # Calculs des statistiques
        date_debut = self.df['Date_debut_Tache'].min()
        date_fin = self.df['Date_fin_tache'].max()
        duree_totale = (date_fin - date_debut).days + 1
        
        # Compter les jours ouvrables
        jours_ouvrables = 0
        date_courante = date_debut
        while date_courante <= date_fin:
            if not self.est_weekend(date_courante):
                jours_ouvrables += 1
            date_courante += timedelta(days=1)
        
        nb_taches = len(self.df)
        duree_moyenne = self.df.apply(lambda x: (x['Date_fin_tache'] - x['Date_debut_Tache']).days + 1, axis=1).mean()
        
        stats_data = [
            ['RÉCAPITULATIF DU PROJET', ''],
            ['', ''],
            ['Trimestre', self.trimestre_choisi],
            ['Nombre de tâches', nb_taches],
            ['', ''],
            ['Date de début', date_debut.strftime('%d/%m/%Y')],
            ['Date de fin', date_fin.strftime('%d/%m/%Y')],
            ['Durée totale (jours calendaires)', duree_totale],
            ['Durée totale (jours ouvrables)', jours_ouvrables],
            ['Durée moyenne par tâche (jours)', f"{duree_moyenne:.1f}"],
            ['', ''],
            ['LISTE DES TÂCHES', ''],
            ['', '']
        ]
        
        # Ajouter les détails des tâches
        for _, tache in self.df.iterrows():
            duree = (tache['Date_fin_tache'] - tache['Date_debut_Tache']).days + 1
            stats_data.append([
                f"#{int(tache['Ordre_exécution'])} - {tache['Intitule_taches']}",
                f"{duree} jours ({tache['Date_debut_Tache'].strftime('%d/%m')} → {tache['Date_fin_tache'].strftime('%d/%m')})"
            ])
        
        stats_df = pd.DataFrame(stats_data, columns=['Catégorie', 'Valeur'])
        stats_df.to_excel(writer, sheet_name='Récapitulatif', index=False)
    
    def appliquer_mise_en_forme(self, fichier, dates):
        """Applique la mise en forme au fichier Excel"""
        wb = load_workbook(fichier)
        
        # Mise en forme de l'onglet Gantt Visuel
        ws = wb['Gantt Visuel']
        
        # Couleurs
        couleur_entete = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        couleur_tache = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        couleur_weekend = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        couleur_info = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        font_entete = Font(bold=True, color="FFFFFF", size=11)
        font_normal = Font(size=10)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Mise en forme de l'en-tête
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(1, col)
            cell.fill = couleur_entete
            cell.font = font_entete
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Identifier les colonnes de dates
        col_debut_dates = 6
        
        # Parcourir toutes les lignes de données
        for row in range(2, ws.max_row + 1):
            # Colonnes d'information
            for col in range(1, col_debut_dates):
                cell = ws.cell(row, col)
                cell.fill = couleur_info
                cell.border = border
                cell.alignment = Alignment(horizontal='center' if col in [1, 5] else 'left', vertical='center')
                cell.font = font_normal
            
            # Colonnes de dates
            for col_idx, date in enumerate(dates, start=col_debut_dates):
                cell = ws.cell(row, col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if self.est_weekend(date):
                    cell.fill = couleur_weekend
                
                if cell.value == '■':
                    cell.fill = couleur_tache
                    cell.font = Font(color="4472C4", size=14, bold=True)
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
        for col in range(col_debut_dates, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 4
        
        ws.freeze_panes = 'F2'
        
        # Ajouter un titre
        ws.insert_rows(1)
        ws.merge_cells('A1:E1')
        titre_cell = ws['A1']
        titre_cell.value = f"DIAGRAMME DE GANTT - {self.trimestre_choisi}"
        titre_cell.font = Font(bold=True, size=14, color="1F4E78")
        titre_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Mise en forme de l'onglet Récapitulatif
        self.formater_recap(wb)
        
        wb.save(fichier)
    
    def formater_recap(self, wb):
        """Formate l'onglet récapitulatif"""
        ws = wb['Récapitulatif']
        
        # Couleurs
        couleur_titre = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        couleur_section = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        couleur_donnees = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Largeurs de colonnes
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 35
        
        # Formater chaque ligne
        for row in range(1, ws.max_row + 1):
            cell_a = ws.cell(row, 1)
            cell_b = ws.cell(row, 2)
            
            # Ligne de titre principal
            if row == 1:
                ws.merge_cells(f'A{row}:B{row}')
                cell_a.fill = couleur_titre
                cell_a.font = Font(bold=True, color="FFFFFF", size=14)
                cell_a.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[row].height = 30
            
            # Lignes de section
            elif 'RÉCAPITULATIF' in str(cell_a.value) or 'LISTE DES TÂCHES' in str(cell_a.value):
                ws.merge_cells(f'A{row}:B{row}')
                cell_a.fill = couleur_section
                cell_a.font = Font(bold=True, color="FFFFFF", size=12)
                cell_a.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[row].height = 25
            
            # Lignes vides
            elif cell_a.value == '' or cell_a.value is None:
                pass
            
            # Lignes de données
            else:
                cell_a.fill = couleur_donnees
                cell_a.font = Font(size=10)
                cell_a.alignment = Alignment(horizontal='left', vertical='center')
                
                cell_b.font = Font(size=10, bold=True)
                cell_b.alignment = Alignment(horizontal='left', vertical='center')
    
    def creer_graphique_gantt(self, fichier):
        """Crée un graphique de Gantt en barres horizontales"""
        wb = load_workbook(fichier)
        ws_gantt = wb['Gantt Visuel']
        
        # Créer une nouvelle feuille pour le graphique
        ws_graphique = wb.create_sheet('Graphique Gantt')
        
        # Préparer les données pour le graphique
        date_debut_projet = self.df['Date_debut_Tache'].min()
        
        # En-têtes
        ws_graphique['A1'] = 'Tâche'
        ws_graphique['B1'] = 'Jours avant début'
        ws_graphique['C1'] = 'Durée (jours)'
        
        # Données
        for idx, (_, tache) in enumerate(self.df.iterrows(), start=2):
            ws_graphique[f'A{idx}'] = f"#{int(tache['Ordre_exécution'])} {tache['Intitule_taches'][:30]}"
            jours_avant = (tache['Date_debut_Tache'] - date_debut_projet).days
            duree = (tache['Date_fin_tache'] - tache['Date_debut_Tache']).days + 1
            ws_graphique[f'B{idx}'] = jours_avant
            ws_graphique[f'C{idx}'] = duree
        
        # Créer le graphique
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = f"Diagramme de Gantt - {self.trimestre_choisi}"
        chart.y_axis.title = 'Tâches'
        chart.x_axis.title = 'Durée (jours)'
        chart.height = 15
        chart.width = 25
        
        # Données du graphique
        data = Reference(ws_graphique, min_col=2, min_row=1, max_row=len(self.df) + 1, max_col=3)
        cats = Reference(ws_graphique, min_col=1, min_row=2, max_row=len(self.df) + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        
        # Inverser l'ordre des tâches pour avoir la tâche 1 en haut
        chart.y_axis.scaling.orientation = "maxMin"
        
        # Couleurs des séries
        from openpyxl.chart.series import DataPoint
        
        # Ajouter le graphique
        ws_graphique.add_chart(chart, "E2")
        
        wb.save(fichier)
    
    def ajouter_legende(self, fichier):
        """Ajoute une légende explicative dans l'onglet Gantt Visuel"""
        wb = load_workbook(fichier)
        ws = wb['Gantt Visuel']
        
        # Trouver la position pour la légende (sous le tableau)
        row_legende = ws.max_row + 3
        
        # Couleurs
        couleur_tache = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        couleur_weekend = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        font_bold = Font(bold=True, size=11)
        font_normal = Font(size=10)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Titre de la légende
        ws.merge_cells(f'A{row_legende}:D{row_legende}')
        cell_titre = ws[f'A{row_legende}']
        cell_titre.value = "LÉGENDE"
        cell_titre.font = Font(bold=True, size=12, color="1F4E78")
        cell_titre.alignment = Alignment(horizontal='center', vertical='center')
        
        # Élément 1 : Tâche active
        row = row_legende + 1
        ws.merge_cells(f'A{row}:B{row}')
        cell_exemple = ws[f'A{row}']
        cell_exemple.value = "■"
        cell_exemple.fill = couleur_tache
        cell_exemple.font = Font(color="4472C4", size=14, bold=True)
        cell_exemple.alignment = Alignment(horizontal='center', vertical='center')
        cell_exemple.border = border
        
        ws.merge_cells(f'C{row}:D{row}')
        cell_desc = ws[f'C{row}']
        cell_desc.value = "Tâche en cours d'exécution"
        cell_desc.font = font_normal
        cell_desc.alignment = Alignment(horizontal='left', vertical='center')
        
        # Élément 2 : Weekend
        row += 1
        ws.merge_cells(f'A{row}:B{row}')
        cell_exemple = ws[f'A{row}']
        cell_exemple.fill = couleur_weekend
        cell_exemple.alignment = Alignment(horizontal='center', vertical='center')
        cell_exemple.border = border
        
        ws.merge_cells(f'C{row}:D{row}')
        cell_desc = ws[f'C{row}']
        cell_desc.value = "Weekend (samedi & dimanche)"
        cell_desc.font = font_normal
        cell_desc.alignment = Alignment(horizontal='left', vertical='center')
        
        # Élément 3 : Informations générales
        row += 2
        ws.merge_cells(f'A{row}:D{row}')
        cell_info = ws[f'A{row}']
        cell_info.value = f"📊 Projet : {self.trimestre_choisi} | Tâches : {len(self.df)} | Période : {self.df['Date_debut_Tache'].min().strftime('%d/%m/%Y')} → {self.df['Date_fin_tache'].max().strftime('%d/%m/%Y')}"
        cell_info.font = Font(size=10, italic=True)
        cell_info.alignment = Alignment(horizontal='center', vertical='center')
        
        wb.save(fichier)

def main():
    print("\n" + "="*50)
    print("   GÉNÉRATEUR DE DIAGRAMME DE GANTT")
    print("         Version Professionnelle")
    print("="*50)
    
    # Chemin du fichier source
    fichier_source = r"D:\ENEM_Working\Apurement salaire\Dossier_travail_Elodie\Fichier_Gant.xlsx"
    
    # Créer l'instance du générateur
    generateur = GanttGenerator(fichier_source)
    
    # Lire les données
    if not generateur.lire_donnees():
        return
    
    # Sélectionner le trimestre
    if not generateur.afficher_menu_trimestre():
        return
    
    # Filtrer les données
    generateur.filtrer_donnees()
    
    # Vérifier les dépendances
    generateur.verifier_dependances()
    
    # Générer le diagramme
    print("\n⏳ Génération du diagramme en cours...")
    print("   📊 Création du diagramme visuel...")
    print("   📈 Création du graphique Gantt...")
    print("   📋 Création du récapitulatif...")
    print("   🎨 Application de la mise en forme...")
    
    generateur.generer_diagramme()
    
    print("\n" + "="*50)
    print("   ✅ TRAITEMENT TERMINÉ AVEC SUCCÈS !")
    print("="*50)
    print("\n📁 Votre fichier contient maintenant :")
    print("   • Onglet 'Gantt Visuel' : Diagramme détaillé avec légende")
    print("   • Onglet 'Graphique Gantt' : Graphique en barres horizontales")
    print("   • Onglet 'Récapitulatif' : Statistiques et liste des tâches")
    print("\n" + "="*50 + "\n")

if __name__ == "__main__":
    main()






